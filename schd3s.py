# ******************************************************************************
# Copyright (c) 2025. All rights reserved.
# # This work is licensed under the Creative Commons Attribution 4.0
# International License. To view a copy of this license,
# visit # http://creativecommons.org/licenses/by/4.0/.
#
# Author: roximn <roximn148@gmail.com>
# ******************************************************************************
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict
from enum import StrEnum, auto

import openpyxl
from shiny import reactive, req
from shiny.express import module, ui, render
import pandas as pd


# ******************************************************************************
class Sections(StrEnum):
    XR = auto()
    US = auto()
    NEURO = auto()
    NON_NEURO_PLAIN = auto()
    NON_NEURO_CONTRAST = auto()


# ******************************************************************************
@module
def modSchedule3s(input, output, session):
    # Radiologists -------------------------------------------------------------
    @reactive.calc
    def radiologistsDf():
        filePath = Path(__file__).parent / 'aaml-doctors.csv'

        try:
            rads = pd.read_csv(filePath)
        except Exception as e:
            print(f"Error reading file: {e}")
            return None

        return rads

    # Radiologists as dict -----------------------------------------------------
    @reactive.calc
    def radChoices() -> dict:
        choices = radiologistsDf()
        if choices is None:
            return {}

        radiologists = dict(zip(choices['ID'], choices['Name']))
        return radiologists

    # Week dates ---------------------------------------------------------------
    @reactive.calc
    def weekDates() -> list[datetime.date]:
        weekStart = req(input.slxWeek())
        return [(weekStart + timedelta(days=i)) for i in range(7)]

    # Excel file ---------------------------------------------------------------
    @reactive.calc
    def excelWorkbook() -> Path:
        weekStart = req(input.slxWeek())
        weekEnd = weekStart + timedelta(days=6)

        xlTitle = 'B2'
        xlDates: list[str] = ['F5:L5', 'F18:L18', 'F25:L25']
        weekLabel = f'{weekStart:%Y%b%d}-{weekEnd:%d}'

        wb = openpyxl.load_workbook(Path(__file__).parent / 'ThreeShiftWeeklyRoster.xlsx')
        ws  = wb['template']
        ws.title = weekLabel
        ws[xlTitle] = f'King Khalid Hospital AlMajmah Radiologist Schedule - {weekStart:%B %Y}'

        # Fill in the dates for each section of the schedule.
        for rng in xlDates:
            for i, cell in enumerate(ws[rng][0]):
                d = weekStart + timedelta(days=i)
                cell.value  = f'{d:%d}'

        # Ensure the parent folder for excelFile exists
        excelFolder = Path(__file__).parent / 'workbooks'
        excelFolder.mkdir(parents=True, exist_ok=True)

        excelFile = excelFolder / f'WeeklyRoster{weekLabel}.xlsx'
        wb.save(excelFile)
        
        ui.notification_show(f"Created {excelFile}", type="message",duration=2)
        
        return excelFile

    # Roster UI ----------------------------------------------------------------
    @render.express
    def rosterUI():
        @render.express
        def weekSelector():
            with ui.layout_columns():
                today = datetime.now()
                lastSunday = today - timedelta(days=(today.isoweekday() % 7))
                ui.input_date('slxWeek', 'Week Selection',
                            value=f'{lastSunday:%Y%m%d}',
                            weekstart=0,
                            daysofweekdisabled=list(range(1,7)))

                ui.input_selectize("slxSection", "Section",
                                    choices={
                                        Sections.XR: 'X-Rays (ER/INP)',
                                        Sections.US: 'Ultrasound (ER/INP/OPD)',
                                        Sections.NEURO: 'Neuro CT (C- and C+)',
                                        Sections.NON_NEURO_PLAIN: 'Non-Neuro CT (C- Only)',
                                        Sections.NON_NEURO_CONTRAST: 'Non-Neuro CT (C+ Only)'},
                                    multiple=False,
                                    remove_button=False)

                ui.input_action_button("btnUpdate", "Update Table")
                ui.input_action_button("btnClear", "Clear")

        @render.express
        def RadiologistSelector():
            rads = req(radChoices())
            dates = weekDates()
            with ui.layout_column_wrap(width=1/8):
                # Dates
                ui.tags.strong('Shifts')
                for i in range(7):
                    ui.tags.div(f'{dates[i]}')

                # Shift C
                ui.tags.strong('C (0000-0800)')
                ui.input_selectize('slxShiftC0', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftC1', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftC2', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftC3', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftC4', label=None, choices=rads, remove_button=True, multiple=False)

                ui.input_selectize('slxShiftC5', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftC6', label=None, choices=rads, remove_button=True, multiple=False)

                # Shift A
                ui.tags.strong('A (0800-1600)')
                ui.input_selectize('slxShiftA0', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftA1', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftA2', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftA3', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftA4', label=None, choices=rads, remove_button=True, multiple=False)

                ui.input_selectize('slxShiftA5', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftA6', label=None, choices=rads, remove_button=True, multiple=False)

                # Shift B
                ui.tags.strong('B (1600-0000)')
                ui.input_selectize('slxShiftB0', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftB1', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftB2', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftB3', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftB4', label=None, choices=rads, remove_button=True, multiple=False)

                ui.input_selectize('slxShiftB5', label=None, choices=rads, remove_button=True, multiple=False)
                ui.input_selectize('slxShiftB6', label=None, choices=rads, remove_button=True, multiple=False)

                # clearSelectors()

    @reactive.effect
    @reactive.event(input.btnClear)
    def clearSelectors():
        selectors = [
            'slxShiftC0', 'slxShiftC1', 'slxShiftC2', 'slxShiftC3', 'slxShiftC4', 'slxShiftC5', 'slxShiftC6',
            'slxShiftA0', 'slxShiftA1', 'slxShiftA2', 'slxShiftA3', 'slxShiftA4', 'slxShiftA5', 'slxShiftA6',
            'slxShiftB0', 'slxShiftB1', 'slxShiftB2', 'slxShiftB3', 'slxShiftB4', 'slxShiftB5', 'slxShiftB6'
        ]
        for selector in selectors:
            ui.update_selectize(selector, selected='')

    @reactive.calc
    def rosterDf():
        """Generates roster data based on the selected shifts and dates"""
        data = [
            ['C', input.slxShiftC0(), input.slxShiftC1(), input.slxShiftC2(), input.slxShiftC3(), input.slxShiftC4(), input.slxShiftC5(), input.slxShiftC6()],
            ['A', input.slxShiftA0(), input.slxShiftA1(), input.slxShiftA2(), input.slxShiftA3(), input.slxShiftA4(), input.slxShiftA5(), input.slxShiftA6()],
            ['B', input.slxShiftB0(), input.slxShiftB1(), input.slxShiftB2(), input.slxShiftB3(), input.slxShiftB4(), input.slxShiftB5(), input.slxShiftB6()],
        ]
        return pd.DataFrame(data, columns=['Shifts', *weekDates()])

    @render.express
    def dataTableUI():
        ui.tags.h3(input.slxSection().upper())

        @render.table(escape=False, classes="table table-striped")
        def _():
            labelMap = radiologistsDf()
            if len(labelMap) <= 0:
                return None

            labels = [f'{n}<br/>({i})' for i, n in zip(labelMap['ID'], labelMap['Abbrev'])]
            labelMap = defaultdict(lambda: '', zip(labelMap['ID'], labels))

            df = rosterDf().copy()
            df.iloc[:, 1:] = df.iloc[:, 1:].apply(lambda x: x.map(labelMap), axis=0)
            return df

    @render.express
    @reactive.event(input.btnUpdate)
    def updateRoster():
        
        labelMap = radiologistsDf()
        if len(labelMap) <= 0:
            print("No radiologists found")
            return None


        labels = [f'{n}\n({i})' for i, n in zip(labelMap['ID'], labelMap['Abbrev'])]
        labelMap = defaultdict(lambda: '', zip(labelMap['ID'], labels))
       
        df = rosterDf().copy()
        df.iloc[:, 1:] = df.iloc[:, 1:].apply(lambda x: x.map(labelMap), axis=0)

        xlSectionRanges: dict[Sections, list[str]] = {
           Sections.XR: ['F26:L26', 'F27:L27', 'F28:L28'],
           Sections.US: ['F19:L19', 'F20:L20', 'F21:L21'],
           Sections.NEURO: ['F6:L6', 'F9:L9', 'F12:L12'],
           Sections.NON_NEURO_PLAIN: ['F7:L7', 'F10:L10', 'F13:L13'],
           Sections.NON_NEURO_CONTRAST: ['F8:L8', 'F11:L11', 'F14:L14'],
        }
        
        xl = excelWorkbook()
        ui.tags.pre(str(xl))
        wb = openpyxl.load_workbook(xl)
        ws = wb.active

        section = input.slxSection()
        for i, rng in enumerate(xlSectionRanges[section]):
            for j, cell in enumerate(ws[rng][0]):
                cell.value  = df.iloc[i, j+1]

        wb.save(xl)
        _ = ui.notification_show(f"Saved {xl.name}", type="message",duration=2)
